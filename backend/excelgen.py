"""Excel workbook generator for the CIBIL Full Detail report.

Reproduces the 5-sheet template (Customer Details, Account Summary, Ownership
Summary, Loan Type Summary, Enquiries) and applies the formatting rules:
red Overdue when > 0, rich-text red for positive DPD tokens in Column H,
DPD-first sequencing, Column K status, and Total rows.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import rules

NAVY = "FF003A8C"
RED = "FFE4002B"
LIGHT = "FFEAF0FA"
GREY = "FFF3F5F9"

HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, color=NAVY, size=13)
TOTAL_FONT = Font(bold=True, color="FF000000", size=10)
TOTAL_FILL = PatternFill("solid", fgColor=GREY)
RED_FONT = Font(color=RED, size=10)
LABEL_FONT = Font(bold=True, color="FF333333", size=10)
MONEY = "#,##0"
THIN = Side(style="thin", color="FFD0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row, headers, widths=None):
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        if widths and c <= len(widths):
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]


def _dpd_rich(raw) -> CellRichText:
    toks = rules.dpd_tokens(raw) or ["000"]
    parts: list = []
    for idx, tok in enumerate(toks):
        if idx:
            parts.append(" | ")
        val = rules.normalize_dpd_token(tok)
        if isinstance(val, int) and val > 0:
            parts.append(TextBlock(InlineFont(color=RED, b=True), tok))
        else:
            parts.append(tok)
    return CellRichText(parts)


# --------------------------------------------------------------------------- #
def _sheet_customer(wb, customer):
    ws = wb.active
    ws.title = "Customer Details"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    ws["A1"] = "Customer Details"
    ws["A1"].font = TITLE_FONT

    d = customer["details"]
    rows = [
        ("Customer Name", d.get("customer_name")),
        ("PAN", d.get("pan")),
        ("CIBIL Score", d.get("cibil_score")),
        ("Date of Birth", d.get("dob")),
        ("Aadhaar / UID", d.get("aadhaar")),
        ("Phone Number(s)", d.get("phone")),
        ("Email ID(s)", d.get("email")),
        ("Report Date", d.get("report_date")),
        ("Source File", customer.get("source_file")),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    _header_row(ws, r, ["Address No.", "Address"])
    r += 1
    for i, addr in enumerate(customer.get("addresses", []), start=1):
        ws.cell(row=r, column=1, value=i).border = BORDER
        ws.cell(row=r, column=2, value=addr).border = BORDER
        r += 1


def _sheet_accounts(wb, data):
    ws = wb.create_sheet("Account Summary")
    ws.sheet_view.showGridLines = False
    widths = [22, 20, 14, 15, 12, 6, 7, 18, 11, 12, 18]
    _header_row(ws, 1, parser_headers(), widths)
    ws.freeze_panes = "A2"

    r = 2
    for a in data["accounts"]:
        ws.cell(row=r, column=1, value=a.get("loan_type")).border = BORDER
        ws.cell(row=r, column=2, value=a.get("financier")).border = BORDER
        c3 = ws.cell(row=r, column=3, value=a.get("sanction"))
        c4 = ws.cell(row=r, column=4, value=a.get("outstanding"))
        c5 = ws.cell(row=r, column=5, value=a.get("overdue"))
        for c in (c3, c4, c5):
            c.number_format = MONEY
            c.border = BORDER
        if (a.get("overdue") or 0) > 0:
            c5.font = RED_FONT
        ws.cell(row=r, column=6, value=a.get("pd")).border = BORDER
        ws.cell(row=r, column=7, value=a.get("ad12m")).border = BORDER
        h = ws.cell(row=r, column=8, value=_dpd_rich(a.get("dpd")))
        h.border = BORDER
        ws.cell(row=r, column=9, value=a.get("ownership")).border = BORDER
        ws.cell(row=r, column=10, value=a.get("date_opened")).border = BORDER
        ws.cell(row=r, column=11, value=a.get("status_derived") or "").border = BORDER
        r += 1

    # Totals row
    total_s = sum((a.get("sanction") or 0) for a in data["accounts"])
    total_o = sum((a.get("outstanding") or 0) for a in data["accounts"])
    total_od = sum((a.get("overdue") or 0) for a in data["accounts"])
    ws.cell(row=r, column=1, value="Totals").font = TOTAL_FONT
    for col, val in ((3, total_s), (4, total_o), (5, total_od)):
        cell = ws.cell(row=r, column=col, value=val)
        cell.number_format = MONEY
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    for col in range(1, 12):
        ws.cell(row=r, column=col).fill = TOTAL_FILL

    # DPD Summary below the table
    r += 2
    ws.cell(row=r, column=1, value="DPD Summary").font = TITLE_FONT
    r += 1
    positive = [a for a in data["accounts"] if a.get("peak_dpd", 0) > 0]
    if not positive:
        ws.cell(row=r, column=1,
                value="No positive DPD observed in the latest reported period.")
    else:
        for a in positive:
            note = (f"{a.get('loan_type')} / {a.get('financier')} "
                    f"({a.get('ownership')}): peak DPD {a.get('peak_dpd')}")
            ws.cell(row=r, column=1, value=note).font = RED_FONT
            r += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _sheet_ownership(wb, data):
    ws = wb.create_sheet("Ownership Summary")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 1, ["Capacity", "Total Sanction Amount",
                        "Total Current Outstanding"], [18, 22, 24])
    r = 2
    for row in data["ownership_summary"]:
        is_total = row.get("total")
        a = ws.cell(row=r, column=1, value=row["capacity"])
        b = ws.cell(row=r, column=2, value=row["sanction"])
        c = ws.cell(row=r, column=3, value=row["outstanding"])
        b.number_format = c.number_format = MONEY
        for cell in (a, b, c):
            cell.border = BORDER
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
        r += 1


def _sheet_loan_type(wb, data):
    ws = wb.create_sheet("Loan Type Summary")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 1, ["Loan Type", "Ownership", "Total No. of Accounts",
                        "No. of Closed Accounts", "Total Sanction Amount",
                        "Total Current Outstanding"], [22, 12, 16, 16, 20, 22])
    r = 2
    for row in data["loan_type_summary"]:
        is_total = row.get("total")
        vals = [row["loan_type"], row["ownership"], row["accounts"],
                row["closed"], row["sanction"], row["outstanding"]]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            if c in (5, 6):
                cell.number_format = MONEY
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
        r += 1


def _sheet_enquiries(wb, data):
    ws = wb.create_sheet("Enquiries")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 18

    ws["A1"] = "Enquiry Summary"
    ws["A1"].font = TITLE_FONT
    _header_row(ws, 2, ["Period", "No. of Enquiries"])
    summ = data["enquiries"]["summary"]
    periods = [("Last 3 Months", "last_3m"), ("Last 6 Months", "last_6m"),
               ("Last 12 Months", "last_12m"), ("Lifetime", "lifetime")]
    r = 3
    for label, key in periods:
        ws.cell(row=r, column=1, value=label).border = BORDER
        ws.cell(row=r, column=2, value=summ.get(key)).border = BORDER
        r += 1

    r += 1
    _header_row(ws, r, ["Member Name", "Enquiry Date", "Enquiry Purpose",
                        "Enquiry Amount"])
    r += 1
    for e in data["enquiries"]["detail"]:
        ws.cell(row=r, column=1, value=e.get("member")).border = BORDER
        ws.cell(row=r, column=2, value=e.get("date")).border = BORDER
        ws.cell(row=r, column=3, value=e.get("purpose")).border = BORDER
        amt = ws.cell(row=r, column=4, value=e.get("amount"))
        amt.number_format = MONEY
        amt.border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Purpose-wise Summary").font = TITLE_FONT
    r += 1
    _header_row(ws, r, ["Enquiry Purpose", "No. of Enquiries", "Total Amount"])
    r += 1
    for row in data["purpose_wise"]:
        is_total = row.get("total")
        a = ws.cell(row=r, column=1, value=row["purpose"])
        b = ws.cell(row=r, column=2, value=row["count"])
        c = ws.cell(row=r, column=3, value=row["amount"])
        c.number_format = MONEY
        for cell in (a, b, c):
            cell.border = BORDER
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
        r += 1


def parser_headers():
    from .parser import ACCOUNT_HEADERS
    return ACCOUNT_HEADERS


def build_workbook(data: dict) -> bytes:
    wb = Workbook()
    customer = dict(data["customer"])
    customer["source_file"] = data.get("source_file")
    _sheet_customer(wb, customer)
    _sheet_accounts(wb, data)
    _sheet_ownership(wb, data)
    _sheet_loan_type(wb, data)
    _sheet_enquiries(wb, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
